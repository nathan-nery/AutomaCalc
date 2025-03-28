from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from interface import *

def activate(titulo, name):
    
    wb = load_workbook(file)

    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    # Hapvida - Colunas que não são hidden: E F P R S U V + W  *** K e M se tiver dengue
    if name == "Hapvida":
        # for k in ('315', '382', '383', '397'):
        for k in ('382', '383', '397'):
            # ws = wb[k]

            # ws.insert_rows(idx=1, amount=1)
            # # ws.delete_cols(1, 1)
            # if k == '315':
            #     ws.delete_cols(1, 1)
            #     ws['A1'] = 'Sinha Junq (CNES 2078791) ' + titulo
            # elif k == '382':
            #     ws.delete_cols(1, 1)
            #     ws['A1'] = 'Diag Joao Pent (CNES 9831665) ' + titulo
                
            # elif k == '383':
            #     ws.delete_cols(1, 1)
            #     ws['A1'] = 'Diag Jd Sumare (CNES 9564284)' + titulo
                
            # elif k == '397':
            #     ws.delete_cols(1, 1)
            #     ws['A1'] = 'São Francisco (CNES 2079275)' + titulo

            ws = wb[k]

            ws.insert_rows(idx=1, amount=1)
            ws.delete_cols(1, 1)
            # if k == '315':
            #     ws['A1'] = 'Sinha Junq (CNES 2078791) ' + titulo
            if k == '382':
                ws['A1'] = 'Diag Joao Pent (CNES 9831665) ' + titulo
            elif k == '383':
                ws['A1'] = 'Diag Jd Sumare (CNES 9564284)' + titulo
            elif k == '397':
                ws['A1'] = 'São Francisco (CNES 2079275)' + titulo
            cell = ws['A1']
            
            
            cell.font = Font(name='Calibri',
                            size='18',
                            bold=True)
            
            ws.row_dimensions[1].height = 30

            for col in range(1,30):
                cell = ws.cell(2,col)
                cell.font = Font(bold=True)
            for c in ('B', 'C', 'D', 'G', 'H', 'I'):
                ws.column_dimensions[c].hidden= True
                color = c + '2'
                ws[color].fill = redFill
            ws.merge_cells('A1:W1')
            ws['F2'] = 'DataNasc'
            ws['A2'] = 'DataSis'
            ws['P2'] = 'Proced'
            ws['Q2'] = 'Result'
            for row in range(1,2500):
                for col in range(1,80):
                    cell = ws.cell(row,col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    if name == "Mais Saúde":
        # for k in ('315', '382', '383', '397'):
            # ws = wb[k]
            ws = wb.active
            ws.insert_rows(idx=1, amount=1)
            ws.delete_cols(1, 1)
            ws['A1'] = titulo + ' - CNES 4828305'
                
            cell = ws['A1']
            
            cell.font = Font(name='Calibri',
                            size='18',
                            bold=True)
            
            ws.row_dimensions[1].height = 30

            for col in range(1,30):
                cell = ws.cell(2,col)
                cell.font = Font(bold=True)
            for c in ('B', 'C', 'D', 'G', 'H', 'I', 'L', 'M', 'O', 'P', 'R', 'U'):
                ws.column_dimensions[c].hidden= True
                color = c + '2'
                ws[color].fill = redFill
            ws.merge_cells('A1:V1')
            ws['V2'] = 'Dist'
            ws['F2'] = 'DataNasc'
            ws['A2'] = 'DataSis'
            ws['P2'] = 'Proced'
            ws['Q2'] = 'Result'
            for row in range(1,2500):
                for col in range(1,80):
                    cell = ws.cell(row,col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    
    # NS1 DANDO PROBLEMA COM HEADER E FOOTER


    # NS1 - Colunas que não são hidden: D E F H L R S T U W X
    # elif name == "NS1":
    #     ws = wb.active
    #     ws.insert_rows(idx=1, amount=1)

    #     ws['A1'] = titulo

    #     cell = ws['A1']
        
    #     cell.font = Font(name='Calibri',
    #                     size='18',
    #                     bold=True)
        
    #     ws.row_dimensions[1].height = 30
    #     # for c in ('A', 'B', 'C', 'G', 'I', 'J', 'K', 'M', 'N', 'O', 'P', 'Q', 'V','Z'):
    #     # Abaixo: Utilizada APENAS quando a sorologia está suspensa! Se há sorologia utilizar a de cima.
    #     # for c in ('A', 'B', 'C', 'G', 'I', 'J', 'K', 'M', 'N', 'O', 'P', 'Q','U', 'V','W', 'X', 'Z'):
        
    #     for c in ('A', 'B', 'C', 'G', 'I', 'J', 'K', 'M', 'N', 'O', 'P', 'Q', 'V','Z'):
    #         ws.column_dimensions[c].hidden= True
    #         color = c + '2'
    #         ws[color].fill = redFill
    #     ws.merge_cells('A1:Y1')
    #     ws['H2'] = 'DataNasc'
    #     ws['A2'] = 'DataSis'
    #     ws['X2'] = 'Num'
    #     ws['P2'] = 'Proced'
    #     for row in range(1,2500):
    #         for col in range(1,80):
    #             cell = ws.cell(row,col)
    #             cell.alignment = Alignment(horizontal='center', vertical='center')

    elif name == "GAL":
        ws = wb.active
        ws.insert_rows(idx=1, amount=1)

        ws['A1'] = titulo

        cell = ws['A1']
        
        cell.font = Font(name='Calibri',
                        size='18',
                        bold=True)
        
        ws.row_dimensions[1].height = 30
        for c in ('C','D','I','J','L','N'):
            ws.column_dimensions[c].hidden= True
        ws.merge_cells('A1:Q1')
        for row in range(1,2500):
            for col in range(1,80):
                cell = ws.cell(row,col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    elif name == "Sabin":
        ws = wb['POSITIVO']

        ws.insert_rows(idx=1, amount=1)
    
        ws['A1'] = titulo

        cell = ws['A1']
        
        cell.font = Font(name='Calibri',
                        size='18',
                        bold=True)
        
        ws.row_dimensions[1].height = 30

        for col in range(1,30):
            cell = ws.cell(2,col)
            cell.font = Font(bold=True)
        for c in ('A', 'G', 'J', 'K', 'I', 'M', 'N'):
            ws.column_dimensions[c].hidden= True
        ws.merge_cells('A1:T1')
        ws['L2'] = "DataNasc"
        for row in range(1,2500):
            for col in range(1,80):
                cell = ws.cell(row,col)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    elif name == "Municipais - Ocultar":
        ws = wb.active
        for c in ('A', 'B', 'C', 'E', 'H', 'J', 'K', 'P'):
            ws.column_dimensions[c].hidden= True                
        for row in range(1,2500):
            for col in range(1,80):
                cell = ws.cell(row,col)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    elif name == "Banco Dengue":
        ws = wb.active
        #'BS', 'BT', --> Sorologia
        #'BY', 'BZ', --> PCR
        #'EM', --> Observação
        for c in ('B', 'C', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BW', 'BX', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CT', 'CU', 'CV', 'CW', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC'):
            ws.column_dimensions[c].hidden= True
            color = c + '1'
            ws[color].fill = redFill
        # ws.merge_cells('A1:FC1')
        ws['A1'] = 'Num Notif'
        ws['D1'] = 'Dt Notif'
        ws['J1'] = 'CNES'
        ws['M1'] = 'Nome'
        ws['N1'] = 'Dt Nasc'
        ws['AA1'] = 'Dist'
        ws['BS1'] = 'Dt Sorol'
        ws['BT1'] = 'Sorol'
        ws['BU1'] = 'Dt Ns1'
        ws['BV1'] = 'NS1'
        ws['BY1'] = 'Dt PCR'
        ws['BZ1'] = 'PCR'
        ws['CR1'] = 'Final'
        ws['CS1'] = 'Crit'
        ws['CX1'] = 'Dt Enc'
        ws['EM1'] = 'Obs'

        # for row in range(1,6500):
        #     for col in range(1,800):
        #         cell = ws.cell(row,col)
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file)
